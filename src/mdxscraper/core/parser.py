from __future__ import annotations

import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

import openpyxl
from chardet import detect

class WordParser:
    """统一的单词解析器接口，提供面向对象的文件解析方式"""

    def __init__(self, file_path: str):
        self.file_path = file_path

    def parse(self) -> List[Dict[str, Any]]:
        """解析文件并返回单词列表，内部处理所有错误"""
        try:
            return self._do_parse()
        except Exception as e:
            logging.error(f"Failed to parse file {self.file_path}: {e}")
            return []

    def _do_parse(self) -> List[Dict[str, Any]]:
        """实际的解析逻辑"""
        ext = Path(self.file_path).suffix.lower()
        if ext not in self._supported_formats:
            raise ValueError(f"Unsupported file format: {ext}")

        return {
            ".xls": self._parse_xls,
            ".xlsx": self._parse_xls,
            ".json": self._parse_json,
            ".txt": self._parse_txt,
            ".md": self._parse_txt,
        }[ext]()

    @property
    def _supported_formats(self):
        """支持的文件格式"""
        return {".xls", ".xlsx", ".json", ".txt", ".md"}

    def _parse_json(self) -> List[Dict[str, Any]]:
        """解析 JSON 文件"""
        return json.load(self._open_encoding_file())

    def _parse_txt(self) -> List[Dict[str, Any]]:
        """解析文本文件（.txt 和 .md）"""
        result = []
        for line in self._open_encoding_file().readlines():
            line = line.strip()
            if len(line) == 0:
                continue
            if line.startswith("#"):
                result.append({"name": line.strip("#").strip(), "words": []})
            else:
                if len(result) == 0:
                    currentTime = datetime.now().strftime("%Y%m%d-%H%M%S")
                    result.append({"name": currentTime, "words": []})
                result[-1]["words"].append(line)
        return result

    def _parse_xls(self) -> List[Dict[str, Any]]:
        """解析 Excel 文件（.xls 和 .xlsx）"""
        wb = openpyxl.load_workbook(self.file_path, read_only=True)
        result = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            words = [row[0].value for row in ws.iter_rows(min_row=ws.min_row, max_row=ws.max_row, max_col=1)]
            words = list(filter(lambda x: x is not None and len(x) > 0, words))
            result.append({"name": sheet_name, "words": words})
        return result

    def _open_encoding_file(self, default_encoding: str = "utf-8"):
        """打开文件并自动检测编码"""
        with open(self.file_path, "rb") as f:
            raw_data = f.read()
        if raw_data.count(b"\n") < 1:
            encoding = default_encoding
        else:
            detection_result = detect(raw_data)
            encoding = detection_result["encoding"]
            confidence = detection_result.get("confidence", 0)
            if confidence < 0.5:
                encoding = default_encoding
        return open(self.file_path, encoding=encoding, errors="ignore")
