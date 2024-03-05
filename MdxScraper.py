#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Module: MdxScraper
Author: VimWei
Created: January 14, 2024
Modified: March 3, 2024

Description:
    Extract specific words from an MDX dictionary and generate HTML, PDF, or JPG files with ease.
    It's an adaptation and upgrade based on the original MdxConverter: https://github.com/noword/MdxConverter
"""

import os
import sys
import time
import json
import shutil
import platform
import tempfile
from pathlib import Path
from collections import OrderedDict
from datetime import datetime, timedelta

import imgkit  # pip install imgkit
import pdfkit  # pip install pdfkit
import openpyxl  # pip install openpyxl
from chardet import detect  # pip install chardet
from base64 import b64encode  # pip install base64
from bs4 import BeautifulSoup  # pip install bs4

from settings import (
    INPUT_PATH, INPUT_NAME,
    DICTIONARY_PATH, DICTIONARY_NAME,
    OUTPUT_PATH, OUTPUT_NAME,
    InvalidAction, INVALID_ACTION, INVALID_WORDS_NAME,
    PDF_OPTIONS, WKHTMLTOPDF_PATHS,
    H1_STYLE, SCRAP_STYLE, ADDITIONAL_STYLES,
)

# import mdict_query
current_script_path = Path(__file__).resolve().parent
path_to_be_added = current_script_path / "lib/mdict-query"
sys.path.append(str(path_to_be_added))
import mdict_query

# Function to open a file with detected encoding
def open_encoding_file(name, default_encoding='utf-8'):
    with open(name, 'rb') as f:
        raw_data = f.read()
    if raw_data.count(b'\n') < 1:
        encoding = default_encoding
    else:
        detection_result = detect(raw_data)
        encoding = detection_result['encoding']
        confidence = detection_result.get('confidence', 0)
        if confidence < 0.5:
            encoding = default_encoding
    return open(name, encoding=encoding, errors='ignore')

# Function to retrieve words from different file types
def get_words(name):
    ext = Path(name).suffix.lower()
    return {'.xls': get_words_from_xls,
            '.xlsx': get_words_from_xls,
            '.json': get_words_from_json,
            '.txt': get_words_from_txt,
            }[ext](name)

# Function to get words from a JSON file
def get_words_from_json(name):
    return json.load(open_encoding_file(name))

# Function to get words from a text file
def get_words_from_txt(name):
    result = []
    for line in open_encoding_file(name).readlines():
        line = line.strip()
        if len(line) == 0:
            continue
        if line.startswith('#'):
            result.append({'name': line.strip('#'), 'words': []})
        else:
            if len(result) == 0:
                currentTime = datetime.now().strftime("%Y%m%d-%H%M%S")
                result.append({'name': currentTime, 'words': []})
            result[-1]['words'].append(line)
    return result

# Function to get words from an Excel file
def get_words_from_xls(name):
    wb = openpyxl.load_workbook(name, read_only=True)
    result = []
    for name in wb.sheetnames:
        ws = wb[name]
        words = [row[0].value for row in ws.iter_rows(min_row=ws.min_row, max_row=ws.max_row, max_col=1)]
        words = list(filter(lambda x: x is not None and len(x) > 0, words))
        result.append({'name': name, 'words': words})
    return result

# Function to retrieve CSS from an MDX dictionary or file
def get_css(soup, mdx_path, dictionary):
    css_name = soup.head.link['href']
    css_path = Path(mdx_path) / css_name
    if css_path.exists():
        css = css_path.read_bytes()
    elif hasattr(dictionary, '_mdd_db'):
        css_key = dictionary.get_mdd_keys('*' + css_name)[0]
        css = dictionary.mdd_lookup(css_key)[0]
    else:
        css = b''
    return css.decode('utf-8')

# Function to merge CSS into the HTML soup
def merge_css(soup, mdx_path, dictionary, append_additinal_styles=True):
    try:
        css = get_css(soup, mdx_path, dictionary)
    except Exception as e:
        return soup
    if append_additinal_styles:
        css += ADDITIONAL_STYLES

    soup.head.link.decompose()
    soup.head.append(soup.new_tag('style', type='text/css'))
    soup.head.style.string = css
    return soup

# Function to determine image format based on file extension
def get_image_format_from_src(src: str) -> str:
    ext = Path(src).suffix.lower()
    if ext == '.png':
        return 'png'
    elif ext in ['.jpg', '.jpeg']:
        return 'jpeg'
    elif ext == '.gif':
        return 'gif'
    elif ext == '.webp':
        return 'webp'
    elif ext == '.svg':
        return 'svg'
    elif ext in ['.tif', '.tiff']:
        return 'tiff'
    elif ext == '.bmp':
        return 'bmp'
    else:
        return 'jpg'

# Function to replace image source with base64 data in HTML soup
def grab_images(soup, dictionary):
    if not hasattr(dictionary, '_mdd_db'):
        return soup

    cache = {}
    for img in soup.find_all('img'):
        if not img.has_attr('src'):
            continue

        src = img['src']
        src_path = src.replace('/', '\\')

        if src_path in cache:
            img['src'] = cache[src_path]
            continue

        lookup_src = src_path
        if not lookup_src.startswith('\\'):
            lookup_src = '\\' + lookup_src

        # Lookup image data
        imgs = dictionary.mdd_lookup(lookup_src)
        if len(imgs) > 0:
            # print(f'Got image {src}')
            image_format = get_image_format_from_src(src)
            base64_str = f'data:image/{image_format};base64,' + b64encode(imgs[0]).decode('ascii')
            cache[src_path] = base64_str
            img['src'] = base64_str

    return soup

# Function to look up a word in the MDX dictionary
def lookup(dictionary, word):
    word = word.strip()
    definitions = dictionary.mdx_lookup(word)
    if len(definitions) == 0:
        definitions = dictionary.mdx_lookup(word, ignorecase=True)
    if len(definitions) == 0:
        definitions = dictionary.mdx_lookup(word.replace('-', ''), ignorecase=True)
    if len(definitions) == 0:
        return ''
    definition = definitions[0]
    if definition.startswith('@@@LINK='):
        return dictionary.mdx_lookup(definition.replace('@@@LINK=', '').strip())[0].strip()
    else:
        return definition.strip()

# Function to convert MDX to HTML
def mdx2html(mdx_file, input_file, output_file, invalid_action=InvalidAction.Collect, with_toc=True):
    print(f'Looking up words in the dictionary and generating HTML output...\n')

    found_count = 0
    not_found_count = 0

    mdx_file = Path(mdx_file)

    dictionary = mdict_query.IndexBuilder(mdx_file)
    lessons = get_words(input_file)

    right_soup = BeautifulSoup('<body style="font-family:Arial Unicode MS;"><div class="right"></div></body>', 'lxml')
    right_soup.find('body').insert_before('\n')
    left_soup = BeautifulSoup('<div class="left"></div>', 'lxml')

    invalid_words = OrderedDict()

    for lesson in lessons:
        # print(lesson['name'])
        h1 = right_soup.new_tag('h1', id='lesson_' + lesson['name'], style=H1_STYLE)
        h1.string = lesson['name']
        right_soup.div.append(h1)

        a = left_soup.new_tag('a', href='#lesson_' + lesson['name'], **{'class': 'lesson'})
        a.string = lesson['name']
        left_soup.div.append(a)
        left_soup.div.append(left_soup.new_tag('br'))
        left_soup.div.append('\n')

        invalid = False
        for word in lesson['words']:
            # print('\t', word)
            result = lookup(dictionary, word)
            if len(result) == 0:  # not found
                not_found_count += 1
                if invalid_action == InvalidAction.Exit:
                    sys.exit()
                elif invalid_action == InvalidAction.Collect:
                    if lesson['name'] in invalid_words:
                        invalid_words[lesson['name']].append(word)
                    else:
                        invalid_words[lesson['name']] = [word, ]
                    continue
                elif invalid_action == InvalidAction.OutputWarning:
                    invalid = True
                    result = f'<div style="padding:0 0 15px 0"><b>WARNING:</b> "{word}" not found</div>'
                else:  # invalid_action == InvalidAction.Collect_OutputWarning
                    if lesson['name'] in invalid_words:
                        invalid_words[lesson['name']].append(word)
                    else:
                        invalid_words[lesson['name']] = [word, ]
                    invalid = True
                    result = f'<div style="padding:0 0 15px 0"><b>WARNING:</b> "{word}" not found</div>'
            else:
                found_count += 1

            definition = BeautifulSoup(result, 'lxml')
            if right_soup.head is None and definition.head is not None:
                right_soup.html.insert_before(definition.head)
                right_soup.head.append(right_soup.new_tag('meta', charset='utf-8'))

            new_div = right_soup.new_tag("div", style=SCRAP_STYLE)
            new_div['id'] = 'word_' + word
            new_div['class'] = 'scrapedword'
            if definition.body:
                new_div.append(definition.body)
            right_soup.div.append('\n')
            right_soup.div.append(new_div)

            a = left_soup.new_tag('a', href='#word_' + word, **{'class': 'word' + (' invalid_word' if invalid else '')})
            invalid = False
            a.string = word
            left_soup.div.append(a)
            left_soup.div.append(left_soup.new_tag('br'))
            left_soup.div.append('\n')

        left_soup.div.append(left_soup.new_tag('br'))

    if with_toc:
        main_div = right_soup.new_tag('div', **{'class': 'main'})
        right_soup.div.wrap(main_div)
        right_soup.div.insert_before(left_soup.div)

    right_soup = merge_css(right_soup, Path(mdx_file).parent, dictionary, with_toc)
    right_soup = grab_images(right_soup, dictionary)

    html = str(right_soup).encode('utf-8')
    html = html.replace(b'<body>', b'').replace(b'</body>', b'', html.count(b'</body>') - 1)
    with open(output_file, "wb") as file:
        file.write(html)

    if len(invalid_words) > 0:
        with open(invalid_words_file, 'w', encoding='utf-8') as fp:
            for lesson, words in invalid_words.items():
                fp.write(f'#{lesson}\n')
                for word in words:
                    fp.write(word + '\n')

    return found_count, not_found_count

# Function to find wkhtmltopdf path
def find_wkhtmltopdf_path():
    if platform.system() == 'Windows':
        return WKHTMLTOPDF_PATHS['Windows']
    elif platform.system() == 'Linux':
        return WKHTMLTOPDF_PATHS['Linux'] if Path(WKHTMLTOPDF_PATHS['Linux']).is_file() else WKHTMLTOPDF_PATHS['Linux_alt']
    elif platform.system() == 'Darwin':
        return WKHTMLTOPDF_PATHS['Darwin']
    else:
        raise ValueError("Unsupported platform")

# Function to convert MDX to PDF
def mdx2pdf(mdx_file, input_file, output_file, invalid_action=InvalidAction.Collect):
    with tempfile.NamedTemporaryFile(suffix='.html', delete=False) as temp:
        TEMP_FILE = temp.name
        result = mdx2html(mdx_file, input_file, TEMP_FILE, invalid_action, False)

    print(f'Converting HTML to PDF...\n')
    config = pdfkit.configuration(wkhtmltopdf=find_wkhtmltopdf_path())
    pdfkit.from_file(TEMP_FILE, output_file, configuration=config, options=PDF_OPTIONS)
    os.remove(TEMP_FILE)
    return result

# Function to convert MDX to JPG
def mdx2jpg(mdx_file, input_file, output_file, invalid_action=InvalidAction.Collect):
    with tempfile.NamedTemporaryFile(suffix='.html', delete=False) as temp:
        TEMP_FILE = temp.name
        result = mdx2html(mdx_file, input_file, TEMP_FILE, invalid_action, False)

    print(f'Converting HTML to JPG...\n')
    imgkit.from_file(TEMP_FILE, output_file, options={'enable-local-file-access': ''})
    os.remove(TEMP_FILE)
    return result

def human_readable_duration(seconds):
    time_delta = timedelta(seconds=seconds)
    hours, remainder = divmod(time_delta.total_seconds(), 3600)
    minutes, int_seconds = divmod(remainder, 60)
    milliseconds = int((seconds - int(hours) * 3600 - int(minutes) * 60 - int(int_seconds)) * 1000)

    parts = []
    if int(hours) > 0:
        parts.append(f'{int(hours):02d} hours')
    if int(minutes) > 0 or int(hours) > 0:
        parts.append(f'{int(minutes):02d} minutes')
    parts.append(f'{int(int_seconds):02d}.{milliseconds:03d} seconds')

    human_readable_time = ''.join(parts)
    return human_readable_time

if __name__ == '__main__':

    print(f'Welcome to MdxScraper：extract specific words from an MDX dictionary and generate HTML, PDF, or JPG with ease！\n')
    start_time = time.time()

    input_file = Path(INPUT_PATH) / INPUT_NAME
    mdx_file = Path(DICTIONARY_PATH) / DICTIONARY_NAME

    currentTime = datetime.now().strftime("%Y%m%d-%H%M%S")
    OUTPUT_PATH = Path(OUTPUT_PATH)
    OUTPUT_PATH.mkdir(parents=True, exist_ok=True)
    if OUTPUT_NAME is None:
        OUTPUT_NAME = currentTime + '_' + Path(INPUT_NAME).stem + '.html'
        output_file = OUTPUT_PATH / OUTPUT_NAME
    else:
        OUTPUT_NAME = currentTime + '_' + OUTPUT_NAME
        output_file = OUTPUT_PATH / OUTPUT_NAME
    invalid_words_file = OUTPUT_PATH / (currentTime + '_' + INVALID_WORDS_NAME)

    backup_input_file = OUTPUT_PATH / (currentTime + '_backup_' + INPUT_NAME)
    shutil.copy(str(input_file), str(backup_input_file))

    output_type = Path(output_file).suffix[1:]
    found, not_found = {
        'html': mdx2html,
        'pdf': mdx2pdf,
        'jpg': mdx2jpg,
    }[output_type](mdx_file, input_file, output_file, INVALID_ACTION)

    if found >0 or INVALID_ACTION in [InvalidAction.OutputWarning, InvalidAction.Collect_OutputWarning]:
        print(f"Success: {found} words extracted from {Path(DICTIONARY_NAME).name}. Refer to {output_file}.\n")
    else:
        print(f"Success: {found} words extracted from {Path(DICTIONARY_NAME).name}.\n")
    if not_found > 0:
        print(f"Failure: {not_found} words not in {Path(DICTIONARY_NAME).name}. Check {invalid_words_file}.\n")
    else:
        print(f"Failure: {not_found} words not in {Path(DICTIONARY_NAME).name}.\n")

    end_time = time.time()
    duration = human_readable_duration(end_time - start_time)
    print(f"The entire process took a total of {duration}.")
